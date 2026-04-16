"""
Submodule to read the data tracks Excel form (.xlsx), extract genome assembly accession number,
and populate the data_tracks.json.

NB! The Excel files cannot contain comments; if it does, pd.read_excel will fail with the error
"This is most probably because the workbook source files contain some invalid XML."

"""

import json
import re
from datetime import datetime
from enum import Enum
from pathlib import Path

import pandas as pd
from add_content_files import TEMPLATE_DIR
from openpyxl import load_workbook

JSON_FILE_NAME = "data_tracks.json"
TEMPLATE_FILE_PATH = TEMPLATE_DIR / JSON_FILE_NAME
INSTRUCTIONS_SHEET_NAME = "Instructions"
INSTRUCTIONS_VERSION_CELL = "A1"
EXPECTED_DATA_TRACKS_FORM_VERSION = "1.3"


class ExpectedExcelColumns(str, Enum):
    """
    Enum for the different expected columns in the Excel file.
    """

    SPECIES_SCIENTIFIC_NAME = "species_scientific_name"
    DATA_TRACK_NAME = "data_track_name"
    DATA_TRACK_DESCRIPTION = "data_track_description"
    assembly_GCA_accession = "assembly_GCA_accession"
    DOI_LINK_TO_REPOSITORY = "doi_link_to_repository"
    FILENAME = "filename"
    PRINCIPAL_INVESTIGATOR_NAME = "principal_investigator_name"
    PRINCIPAL_INVESTIGATOR_AFFILIATION = "principal_investigator_affiliation"
    BUSCO_STATS = "BUSCO_stats"
    DIRECT_LINK_TO_FILE_FOR_DOWNLOAD = "direct_link_to_file_for_download"
    DOI_LINK_TO_SCIENTIFIC_ARTICLE = "doi_link_to_scientific_article"


EXPECTED_EXCEL_COLUMNS = {column.value for column in ExpectedExcelColumns}

# Backwards compatible fallback for older form versions that did not have the BUSCO_stats column
OPTIONAL_EXCEL_COLUMNS = {ExpectedExcelColumns.BUSCO_STATS.value}
REQUIRED_EXCEL_COLUMNS = EXPECTED_EXCEL_COLUMNS - OPTIONAL_EXCEL_COLUMNS


def validate_excel_form_version(
    spreadsheet_file_path: str, expected_version: str = EXPECTED_DATA_TRACKS_FORM_VERSION
) -> None:
    """
    Validate Excel form version from Instructions!A1.
    Expected format: "Instructions (version X.Y)"
    """
    try:
        workbook = load_workbook(spreadsheet_file_path, read_only=True, data_only=True)
    except ValueError as e:
        raise ValueError("Your spreadsheet likely contains comments/invalid XML. Remove comments and re-run.") from e
    if INSTRUCTIONS_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Missing '{INSTRUCTIONS_SHEET_NAME}' sheet in data tracks form. " f"Expected version {expected_version}."
        )

    instructions_sheet = workbook[INSTRUCTIONS_SHEET_NAME]
    cell_value = instructions_sheet[INSTRUCTIONS_VERSION_CELL].value
    version_text = str(cell_value or "").strip()
    version_match = re.search(r"Instructions\s*\(version\s*([0-9]+(?:\.[0-9]+)?)\)", version_text, flags=re.I)

    if not version_match:
        raise ValueError(
            f"Could not detect data tracks form version from "
            f"{INSTRUCTIONS_SHEET_NAME}!{INSTRUCTIONS_VERSION_CELL}. Expected 'Instructions (version {expected_version})'."
        )

    detected_version = version_match.group(1)
    if detected_version != expected_version:
        raise ValueError(
            f"Unsupported data tracks form version {detected_version}. "
            f"Expected version {expected_version}. Please use the latest template."
        )


def validate_excel_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Validate that spreadsheet columns match the enum-defined schema exactly.
    Users should not add or alter any column names.
    """
    actual_columns = {str(col) for col in df.columns}

    missing_columns = sorted(REQUIRED_EXCEL_COLUMNS - actual_columns)
    unexpected_columns = sorted(actual_columns - EXPECTED_EXCEL_COLUMNS)
    if missing_columns or unexpected_columns:
        raise ValueError(
            "Invalid columns in data tracks sheet. "
            f"Missing: {missing_columns or 'None'}. "
            f"Unexpected: {unexpected_columns or 'None'}. "
            "Only enum-defined columns are allowed."
        )

    return df


def df_row_to_json(row: pd.Series, template_json: str) -> dict[str, str]:
    """
    Convert a row of the DataFrame to a JSON object using the JSON template.
    Handle missing values by checking if the value is None or NaN: if so,
    use the default value from the template instead. Columns that are present
    in the template but not in the row will be passed through as is.
    """

    data_track = json.loads(template_json)
    accession_or_doi = None

    if "data_track_name" in row and pd.notna(row["data_track_name"]):
        data_track["dataTrackName"] = row["data_track_name"]
    if "assembly_GCA_accession" in row and pd.notna(row["assembly_GCA_accession"]):
        data_track["assemblyGCAAccession"] = row["assembly_GCA_accession"]
        if data_track.get("dataTrackName") == "Genome":
            accession_or_doi = row["assembly_GCA_accession"]
    if "data_track_description" in row and pd.notna(row["data_track_description"]):
        data_track["description"] = row["data_track_description"]
    if "doi_link_to_repository" in row and pd.notna(row["doi_link_to_repository"]):
        data_track["links"][1]["Website"] = row["doi_link_to_repository"]
    if "filename" in row and pd.notna(row["filename"]):
        data_track["fileName"] = row["filename"]
    if "principal_investigator_name" in row and pd.notna(row["principal_investigator_name"]):
        data_track["principalInvestigator"] = row["principal_investigator_name"]
    if "principal_investigator_affiliation" in row and pd.notna(row["principal_investigator_affiliation"]):
        data_track["principalInvestigatorAffiliation"] = row["principal_investigator_affiliation"]
    if "BUSCO_stats" in row and pd.notna(row["BUSCO_stats"]):
        data_track["buscoStats"] = row["BUSCO_stats"]
    if "direct_link_to_file_for_download" in row and pd.notna(row["direct_link_to_file_for_download"]):
        data_track["links"][0]["Download"] = row["direct_link_to_file_for_download"]
    if "doi_link_to_scientific_article" in row and pd.notna(row["doi_link_to_scientific_article"]):
        data_track["links"][2]["Article"] = row["doi_link_to_scientific_article"]
    if "firstDateOnPortal" in row and pd.notna(row["firstDateOnPortal"]):
        data_track["firstDateOnPortal"] = row["firstDateOnPortal"]
    else:
        data_track["firstDateOnPortal"] = datetime.now().strftime("%d %B %Y")

    if accession_or_doi:
        data_track["accessionOrDOI"] = accession_or_doi

    return data_track


def parse_excel_file(spreadsheet_file_path: str, sheet_name: str) -> list[dict]:
    """
    Parse the Excel file with Pandas and return a JSON-style structure (list of dicts).
    JSON was chosen over dataclass since the data is used to populate data_track.json.
    """
    validate_excel_form_version(spreadsheet_file_path=spreadsheet_file_path)
    try:
        df = pd.read_excel(spreadsheet_file_path, sheet_name=sheet_name, engine="openpyxl")
    except ValueError as e:
        raise ValueError("Your spreadsheet likely contains comments/invalid XML. Remove comments and re-run.") from e
    df = validate_excel_columns(df)

    with open(TEMPLATE_FILE_PATH, "r") as file:
        template_json = json.dumps(json.load(file)[0])

    user_data_tracks = [df_row_to_json(row, template_json) for _, row in df.iterrows()]

    return user_data_tracks


def populate_data_tracks_json(user_data_tracks: list[dict], assets_dir_path: Path) -> None:
    """
    Write the data tracks list of dictionaries to a JSON file.
    """
    output_json_path = assets_dir_path / JSON_FILE_NAME

    with open(output_json_path, "w") as json_file:
        json.dump(user_data_tracks, json_file, indent=2)
        print(f"File created: {output_json_path.resolve()}")
